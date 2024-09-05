from tkinter import *
from tkinter.ttk import Combobox
import openpyxl
from openpyxl import Workbook
from tkinter.messagebox import showerror
import pathlib

root = Tk()
root.geometry('600x400')
root.title('Account Details')
root.config(bg='#D3D3D3')
root.resizable(width=0, height=0)

file = pathlib.Path('Success_Bank.xlsx')

if not file.exists():
    file = Workbook()
    sheet = file.active
    sheet['A1'] = 'Account Name'
    sheet['B1'] = 'Account Number'
    sheet['C1'] = 'Account Type'
    sheet['D1'] = 'Account Balance'
    file.save('Success_Bank.xlsx')

def Submit():
    name = name1.get()
    acc_num=ent2.get()
    acc_type = ent3.get()
    acc_balance = amount1.get()

    if not name or not acc_type or not acc_balance:
        showerror("Input Error", "Please fill all fields.")
        return

    file = openpyxl.load_workbook('Success_Bank.xlsx')
    sheet = file.active
    sheet.cell(column=1, row=sheet.max_row + 1, value=name)
    sheet.cell(column=2, row=sheet.max_row, value=acc_num)
    sheet.cell(column=3, row=sheet.max_row, value=acc_type)
    sheet.cell(column=4, row=sheet.max_row, value=acc_balance)
    file.save(r'Success_Bank.xlsx')

    ent1.delete(0, END)
    ent2.delete(0, END)
    ent3.set('')
    ent4.delete(0, END)
    ent5.delete(0, END)


class Account:
    def __init__(self, account_name, account_type, balance=0):
        self.account_name = account_name
        self.account_type = account_type
        self._balance = balance

    def deposit(self, amount):
        if amount > 0:
            self._balance += amount
    

    def withdraw(self, amount):
        if amount > 0 and amount <= self._balance:
            self._balance -= amount

    def get_balance(self):
        return self._balance

account_instance = None


def Deposit():
    global account_instance
    try:
        name = name1.get()
        acc_type = ent3.get()
        amount = int(amount1.get())

        if account_instance is None:
            account_instance = Account(name, acc_type)
        account_instance.deposit(amount)

        result = f'Your total balance is: {account_instance.get_balance()}'
        result_label.config(text=result)

    except ValueError:
        showerror('Invalid Input', 'Please enter a valid amount')
        result_label.config(text='')

def Withdraw():
    global account_instance
    try:
        amount = int(withdraw_amount.get())

        if account_instance is None:
            showerror('No Account', 'Please create an account and deposit money first.')
            return

        account_instance.withdraw(amount)
        result = f'Your total balance is: {account_instance.get_balance()}'
        result_label.config(text=result)

    except ValueError:
        showerror('Invalid Input', 'Please enter a valid withdrawal amount')
        result_label.config(text='')

l1 = Label(root, text='Account Name:', font=('calibri', 12), bg='#D3D3D3')
l1.place(x=10, y=30)

l2 = Label(root, text='Account Number:', font=('calibri', 12), bg='#D3D3D3')
l2.place(x=10, y=60)

l3 = Label(root, text='Account Type:', font=('calibri', 12), bg='#D3D3D3')
l3.place(x=10, y=90)

l4 = Label(root, text='Deposit Amount:', font=('calibri', 12), bg='#D3D3D3')
l4.place(x=10, y=120)

l5 = Label(root, text='Withdraw Amount:', font=('calibri', 12), bg='#D3D3D3')
l5.place(x=10, y=150)

name1 = StringVar()
accnumb = StringVar()
amount1 = StringVar()
withdraw_amount = StringVar()

ent1 = Entry(root, textvariable=name1, font=('Calibri', 12), width=41, bd=0.2)
ent1.place(x=130, y=35)

ent2 = Entry(root, textvariable=accnumb, font=('Calibri', 12), width=41, bd=0.2)
ent2.place(x=130, y=65)

ent3 = Combobox(root, values=['Savings Account', 'Current Account'], font=('Calibri', 12), width=38)
ent3.place(x=130, y=90)

ent4 = Entry(root, textvariable=amount1, font=('Calibri', 12), width=41, bd=0.2)
ent4.place(x=130, y=125)

ent5 = Entry(root, textvariable=withdraw_amount, font=('Calibri', 12), width=41, bd=0.2)
ent5.place(x=130, y=155)

bt1 = Button(root, text='Deposit', command=Deposit, font=('Calibri', 12), width=20, bd=0.2)
bt1.place(x=130, y=250)

bt2 = Button(root, text='Withdraw', command=Withdraw, font=('Calibri', 12), width=20, bd=0.2)
bt2.place(x=130, y=300)

bt3 = Button(root, text='Submit', command=Submit, font=('Calibri', 12), width=20, bd=0.2)
bt3.place(x=300, y=300)

result_label = Label(root, text='', font=('calibri', 12), bg='#D3D3D3')
result_label.place(x=130, y=350)

root.mainloop()
