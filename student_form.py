from tkinter import *
from tkinter.ttk import Combobox
import openpyxl
from openpyxl import Workbook
import pathlib

root = Tk()
root.geometry('500x400')
root.title('Student Form')
root.config(bg='#E0EEE0')
root.resizable(width=0,height=0)

file=pathlib.Path('Student_Form.xlsx')

if file.exists():
    pass
else:
    file = Workbook ()
    sheet = file.active
    sheet['A1'] = 'Name'
    sheet['B1'] = 'Age'
    sheet['C1'] = 'Contact'
    sheet['D1'] = 'Gender'
    sheet['E1'] = 'Address'
    sheet['F1'] = 'Program'
    sheet['G1'] = 'Email'
    file.save('Student_Form.xlsx')

def Submit():
    name = name1.get()
    age = age1.get()
    contact=contact1.get()
    gender=ent4.get()
    address=ent5.get(1.0, END)
    program=prog1.get()
    email=email1.get()

    file= openpyxl.load_workbook('Student_Form.xlsx')
    sheet = file.active
    sheet.cell(column=1, row=sheet.max_row+1, value=name)
    sheet.cell(column=2, row=sheet.max_row, value=age)
    sheet.cell(column=3, row=sheet.max_row, value=contact)
    sheet.cell(column=4, row=sheet.max_row, value=gender)
    sheet.cell(column=5, row=sheet.max_row, value=address)
    sheet.cell(column=6, row=sheet.max_row, value=program)
    sheet.cell(column=7, row=sheet.max_row, value=email)
    file.save(r'Student_Form.xlsx')

    ent1.delete(0, END)
    ent2.delete(0, END)
    ent3.delete(0, END)
    ent4.delete(0, END)
    ent5.delete(1.0, END)
    ent6.delete(0, END)
    ent7.delete(0, END)
    ent1.focus()



def Clear():
    ent1.delete(0, END)
    ent2.delete(0, END)
    ent3.delete(0, END)
    ent5.delete(1.0, END)
    ent6.delete(0, END)
    ent7.delete(0, END)
    ent1.focus()

l1 = Label(root, text='Name', font=('Calibri, bold',13),bg='#E0EEE0')
l1.place(x=10, y=30)

l2 = Label(root, text='Age', font=('Calibri, bold',13),bg='#E0EEE0')
l2.place(x=10, y=60)

l3 = Label(root, text='Contact', font=('Calibri, bold',13),bg='#E0EEE0')
l3.place(x=10, y=90)

l4 = Label(root, text='Gender', font=('Calibri, bold',13),bg='#E0EEE0')
l4.place(x=270, y=60)

l5 = Label(root, text='Address', font=('Calibri, bold',13),bg='#E0EEE0')
l5.place(x=10, y=150)

l6 = Label(root, text='Email', font=('Calibri, bold',13),bg='#E0EEE0')
l6.place(x=10, y=260)

l7 = Label(root, text='Program of Study', font=('Calibri, bold',13),bg='#E0EEE0')
l7.place(x=10, y=220)

name1=StringVar()
age1=StringVar()
contact1=StringVar()
address1=StringVar()
email1=StringVar()
prog1=StringVar()


ent1=Entry(root, textvariable=name1,font=('Calibri, bold',13), width=41,bd=0.5)
ent1.place(x=80,y=30)

ent2=Entry(root, textvariable=age1,font=('Calibri, bold',13), width=10,bd=0.5)
ent2.place(x=80,y=60)

ent3=Entry(root, textvariable=contact1,font=('Calibri, bold',13), width=41,bd=0.5)
ent3.place(x=80,y=90)

ent4=Combobox(root, values=['Male','Female'],font=('Calibri, bold',13), width=10)
ent4.place(x=340,y=60)

ent5=Text(root, width=41, height=4,bd=0.5, font=('Calibri, bold',13))
ent5.place(x=80,y=120)

ent6=Entry(root, textvariable=email1,font=('Calibri, bold',13), width=41,bd=0.5)
ent6.place(x=80,y=260)

ent7=Entry(root, textvariable=prog1,font=('Calibri, bold',13), width=41,bd=0.5)
ent7.place(x=80,y=220)


bt1=Button(root, text='Submit',command=Submit, font=('Calibri, bold',13))
bt1. place(x=150, y=330)

bt2=Button(root, text='Clear',command=Clear, font=('Calibri, bold',13))
bt2. place(x=230, y=330)


root.mainloop()

